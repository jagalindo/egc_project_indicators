#!/bin/python
import git
from typing import List
import openpyxl
import gitfame
import os.path
from os import path
class User:
    def __init__(self, name:str, tokens:List[str]):
        self.name = name
        self.tokens = tokens  
        self.loc = 0
        self.commits = 0

    def __str__(self):
        return f"Name: {self.name}, Tokens: {self.tokens}"
    
    def __html__(self):
        pass
        
class Project:
    def __init__(self, name:str, url:str, branches:List[str], users:List[User]):
        self.name = name
        self.url = url
        self.branches = branches
        self.users = users
        self.total_loc = 0
        self.total_commits = 0

    def __str__(self):
        return f"Name: {self.name}, Url: {self.url}, Branches: {self.branches}, Users: {self.users}"

    def __html__(self):
        pass

def main():
    projects = get_projects("Libro1.xlsx")
    for project in projects:
        process_project(project)
    
    

def get_projects(excel_file:str) -> List[Project]:
    projects=[]
    # load excel with its path
    wrkbk = openpyxl.load_workbook(excel_file)
    sh = wrkbk.active
    
    # iterate through excel and display data
    for i in range(2,3):# sh.max_row+1):
        project=Project(sh.cell(row=i, column=1).value.strip(),sh.cell(row=i, column=2).value.strip(),[item.strip() for item in sh.cell(row=i, column=3).value.split(",")],[])
        users=sh.cell(row=i, column=4).value.split("\n")
        for user in users:
            u = User(user.split("->")[0].strip(),[item.strip for item in user.split("->")[1].split(",")])
            project.users.append(u)
        projects.append(project)
    return projects

def process_project(project:Project):
    if not path.exists(project.name):
        repo=git.Repo.clone_from(project.url, project.name, 
            branch=project.branches[0])
    res=gitfame.main(['--sort=commits', '-wt', project.name])
   
    print(res)

if __name__ == "__main__":
    main()